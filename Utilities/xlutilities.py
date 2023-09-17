import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(xlpath, sheetname):
    wb = openpyxl.load_workbook(xlpath)
    sheet = wb[sheetname]
    return sheet.max_row


def getColoumnCount(xlpath, sheetname):
    wb = openpyxl.load_workbook(xlpath)
    sheet = wb[sheetname]
    return sheet.max_column


def readData(xlpath, sheetname, r_no, c_no):
    wb = openpyxl.load_workbook(xlpath)
    sheet = wb[sheetname]
    return sheet.cell(r_no, c_no).value


def writeData(xlpath, sheetname, r_no, c_no, data):
    wb = openpyxl.load_workbook(xlpath)
    sheet = wb[sheetname]
    sheet.cell(r_no, c_no).value = data
    wb.save(xlpath)


def fillGreenColour(xlpath, sheetname, r_no, c_no):
    wb = openpyxl.load_workbook(xlpath)
    sheet = wb[sheetname]
    green = PatternFill(start_color="60b212",
                        end_color="60b212",
                        fill_type="solid")
    sheet.cell(r_no, c_no).fill = green
    wb.save(xlpath)


def fillRedColour(xlpath, sheetname, r_no, c_no):
    wb = openpyxl.load_workbook(xlpath)
    sheet = wb[sheetname]
    red = PatternFill(start_color="ff0000",
                      end_color="ff0000",
                      fill_type="solid")
    sheet.cell(r_no, c_no).fill = red
    wb.save(xlpath)
